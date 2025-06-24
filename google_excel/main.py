import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import openpyxl
from openpyxl.styles import numbers
import os
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
import re
import json
import sys
sys.stdout.reconfigure(encoding='utf-8')

# Настройка авторизации для Google Sheets и Google Drive API
SCOPE = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']
# Укажите путь к вашему JSON-файлу с учетными данными
CREDS_FILE = 'google_excel\credentials.json'


def authenticate():
    creds = ServiceAccountCredentials.from_json_keyfile_name(CREDS_FILE, SCOPE)
    client = gspread.authorize(creds)
    drive_service = build('drive', 'v3', credentials=creds)
    return client, drive_service


def get_excel_files_in_folder(folder_url):
    try:
        folder_id = folder_url.split('/')[-1].split('?')[0]
        print(f"Extracted Folder ID: {folder_id}")

        client, drive_service = authenticate()
        query = f"'{folder_id}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
        results = drive_service.files().list(
            q=query, fields="nextPageToken, files(id, name, mimeType)").execute()
        files = results.get('files', [])
        print(f"Found {len(files)} Excel files: {[f['name'] for f in files]}")

        return files, drive_service
    except Exception as e:
        print(f"Error in get_excel_files_in_folder: {e}")
        return [], None


def download_excel_file(drive_service, file_id, file_name):
    try:
        request = drive_service.files().get_media(fileId=file_id)
        file_path = f"temp_{file_name}"
        fh = io.FileIO(file_path, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        print(f"Downloaded: {file_name}")
        return file_path
    except Exception as e:
        print(f"Error downloading {file_name}: {e}")
        return None


def process_excel_file(file_path):
    try:
        # Читаем только вычисленные значения
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        data = [[cell.value if cell.value is not None else '' for cell in row]
                for row in sheet.iter_rows()]

        result_data = []
        current_model = ''
        header_row = None
        col_indices = {}
        is_discount_opt = False
        is_discount_rozn = False

        # Определяем ключевые слова для заголовков столбцов
        keywords = {
            'model': ['модель:', 'назва моделі'],
            'size': ['розмір'],
            'price_opt': ['оптова ціна', 'ціна гурт'],
            'price_opt_akciya': ['акційна оптова ціна', 'акційна гуртова ціна', 'оптова ціна акція'],
            'price_rozn': ['роздрібна ціна', 'ціна роздріб'],
            'price_rozn_akciya': ['акційна роздрібна ціна', 'акційна роздрібна ціна', 'роздрібна ціна акція']
        }

        i = 0
        while i < len(data):
            row = data[i]
            # Пропускаем пустые строки
            if not any(row):
                i += 1
                continue
            # Ищем строку с моделью
            model_found = False
            for cell in row:
                if not isinstance(cell, str):
                    continue
                cell_lower = cell.lower().strip()
                for variant in keywords['model']:
                    if cell_lower.startswith(variant):
                        current_model = cell.replace(
                            variant, '').strip().strip('"')
                        print(f"Found model: {current_model}")
                        model_found = True
                        i += 1
                        break
                if model_found:
                    break
            if model_found:
                # Ищем заголовок таблицы
                col_indices = {}
                is_discount_opt = False
                is_discount_rozn = False
                while i < len(data):
                    row = data[i]
                    contains_size = False
                    for cell in row:
                        if not isinstance(cell, str):
                            continue
                        if any(cell.lower().startswith(v) for v in keywords['size']):
                            contains_size = True
                            break
                    if contains_size:
                        header_row = row
                        # Определяем индексы столбцов
                        for j, cell in enumerate(header_row):
                            if not isinstance(cell, str):
                                continue
                            cell_lower = cell.lower().strip()
                            for key, variants in keywords.items():
                                if key == 'model':
                                    continue
                                for variant in variants:
                                    if cell_lower.startswith(variant.lower()):
                                        col_indices[key] = j
                                        # Проверяем, содержит ли заголовок скидку
                                        if key == 'price_opt_akciya' and '% ' in cell_lower:
                                            is_discount_opt = True
                                        if key == 'price_rozn_akciya' and '% ' in cell_lower:
                                            is_discount_rozn = True
                                        break
                        print(
                            f"Found header with columns: {col_indices}, discount_opt: {is_discount_opt}, discount_rozn: {is_discount_rozn}")
                        i += 1
                        break
                    i += 1

                # Обрабатываем строки данных
                while i < len(data):
                    row = data[i]
                    if not any(row):  # Пропускаем пустые строки
                        i += 1
                        continue
                    # Проверяем, является ли строка началом новой модели
                    is_new_model = False
                    for cell in row:
                        if not isinstance(cell, str):
                            continue
                        cell_lower = cell.lower().strip()
                        if any(cell_lower.startswith(v) for v in keywords['model']):
                            is_new_model = True
                            i -= 1  # Вернемся к этой строке для обработки новой модели
                            break
                    if is_new_model:
                        break
                    # Проверяем, содержит ли строка размер
                    size_value = row[col_indices.get(
                        'size', -1)] if 'size' in col_indices else ''
                    if not isinstance(size_value, str) or not size_value.strip():
                        i += 1
                        continue
                    # Пропускаем строки с описаниями
                    if any(cell.lower().startswith('основні характеристики') for cell in row if isinstance(cell, str)):
                        i += 1
                        continue
                    # Извлекаем данные
                    row_dict = {
                        'Model_rez': current_model,
                        'Size_rez': str(size_value).strip()
                    }
                    # Оптовая цена
                    price_opt = row[col_indices.get(
                        'price_opt', -1)] if 'price_opt' in col_indices else ''
                    if isinstance(price_opt, (int, float)):
                        row_dict['Price_Opt'] = round(float(price_opt), 2)
                    else:
                        row_dict['Price_Opt'] = ''
                    # Акционная оптовая цена
                    price_opt_akciya = row[col_indices.get(
                        'price_opt_akciya', -1)] if 'price_opt_akciya' in col_indices else ''
                    if isinstance(price_opt_akciya, str) and price_opt_akciya.lower().strip() == 'відсутня':
                        row_dict['Price_Opt_Akciya'] = ''
                    elif isinstance(price_opt_akciya, (int, float)):
                        row_dict['Price_Opt_Akciya'] = round(
                            float(price_opt_akciya), 2)
                    elif is_discount_opt and price_opt and isinstance(price_opt, (int, float)):
                        row_dict['Price_Opt_Akciya'] = round(
                            float(price_opt) * 0.7, 2)
                    else:
                        row_dict['Price_Opt_Akciya'] = ''
                    # Розничная цена
                    price_rozn = row[col_indices.get(
                        'price_rozn', -1)] if 'price_rozn' in col_indices else ''
                    if isinstance(price_rozn, (int, float)):
                        row_dict['Price_Rozn'] = round(float(price_rozn), 2)
                    else:
                        row_dict['Price_Rozn'] = ''
                    # Акционная розничная цена
                    price_rozn_akciya = row[col_indices.get(
                        'price_rozn_akciya', -1)] if 'price_rozn_akciya' in col_indices else ''
                    if isinstance(price_rozn_akciya, str) and price_rozn_akciya.lower().strip() == 'відсутня':
                        row_dict['Price_Rozn_Akcia'] = ''
                    elif isinstance(price_rozn_akciya, (int, float)):
                        row_dict['Price_Rozn_Akcia'] = round(
                            float(price_rozn_akciya), 2)
                    elif is_discount_rozn and price_rozn and isinstance(price_rozn, (int, float)):
                        row_dict['Price_Rozn_Akcia'] = round(
                            float(price_rozn) * 0.7, 2)
                    else:
                        row_dict['Price_Rozn_Akcia'] = ''
                    result_data.append(row_dict)
                    i += 1
            else:
                i += 1

        return result_data
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return []


def main(folder_url):
    excel_files, drive_service = get_excel_files_in_folder(folder_url)
    all_data = []

    for file in excel_files:
        print(f"Processing file: {file['name']}")
        file_path = download_excel_file(
            drive_service, file['id'], file['name'])
        if file_path:
            data = process_excel_file(file_path)
            all_data.extend(data)
            try:
                os.remove(file_path)
                print(f"Deleted temporary file: {file_path}")
            except Exception as e:
                print(f"Error deleting {file_path}: {e}")

    if all_data:
        # Создаем новый Excel-файл с форматированием
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Записываем заголовки
        headers = ['Model_rez', 'Size_rez', 'Price_Opt',
                   'Price_Opt_Akciya', 'Price_Rozn', 'Price_Rozn_Akcia']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx).value = header

        # Записываем данные
        for row_idx, row_data in enumerate(all_data, start=2):
            ws.cell(row=row_idx, column=1).value = row_data['Model_rez']
            ws.cell(row=row_idx, column=2).value = row_data['Size_rez']
            for col_idx, key in enumerate(['Price_Opt', 'Price_Opt_Akciya', 'Price_Rozn', 'Price_Rozn_Akcia'], start=3):
                value = row_data[key]
                cell = ws.cell(row=row_idx, column=col_idx)
                if value and isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00'  # Формат с запятой и двумя знаками
                else:
                    cell.value = value

        output_file = 'result.xls'
        wb.save(output_file)
        print(f"Result saved to {output_file}")
    else:
        print("No data to save.")


if __name__ == '__main__':
    # Замените на вашу ссылку

    with open("secret_info.json") as f:
        secrets = json.load(f)

    api_key = secrets["Path_Google"]

    FOLDER_URL = ''
    main(FOLDER_URL)
